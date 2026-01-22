"""
🧾 OmniAI Fiscal - Exclusão ICMS-ST
Aplicação Streamlit para cálculo de créditos PIS/COFINS

Execute com: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import json
import io
import zipfile
import tempfile
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass, asdict
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, List, Optional, Generator, Tuple
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# =============================================================================
# CONFIGURAÇÃO DA PÁGINA
# =============================================================================

st.set_page_config(
    page_title="OmniAI Fiscal - Exclusão ICMS-ST",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS Customizado
st.markdown("""
<style>
    /* Tema geral */
    .main {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
    }
    
    /* Header */
    .header-container {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a87 100%);
        padding: 2rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 10px 40px rgba(30, 58, 95, 0.3);
    }
    
    .header-title {
        color: white;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
    }
    
    .header-subtitle {
        color: #94a3b8;
        font-size: 1.1rem;
        margin-top: 0.5rem;
    }
    
    /* Cards de métricas */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border-left: 4px solid #10b981;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(0,0,0,0.12);
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #10b981;
        margin: 0;
    }
    
    .metric-label {
        color: #64748b;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    /* Upload area */
    .upload-area {
        border: 2px dashed #cbd5e1;
        border-radius: 12px;
        padding: 2rem;
        text-align: center;
        background: #f8fafc;
        transition: all 0.3s;
    }
    
    .upload-area:hover {
        border-color: #3b82f6;
        background: #eff6ff;
    }
    
    /* Tabela de resultados */
    .results-table {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
    }
    
    /* Botões */
    .stButton>button {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        border: none;
        padding: 0.75rem 2rem;
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s;
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(16, 185, 129, 0.4);
    }
    
    /* Progress bar */
    .stProgress > div > div {
        background: linear-gradient(90deg, #10b981, #3b82f6);
    }
    
    /* Sidebar */
    .css-1d391kg {
        background: #1e293b;
    }
    
    /* Info boxes */
    .info-box {
        background: #eff6ff;
        border: 1px solid #bfdbfe;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .warning-box {
        background: #fef3c7;
        border: 1px solid #fcd34d;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .success-box {
        background: #d1fae5;
        border: 1px solid #6ee7b7;
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Custom scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: #f1f5f9;
    }
    
    ::-webkit-scrollbar-thumb {
        background: #94a3b8;
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: #64748b;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class SpedHeader:
    cod_ver: str = ""
    tipo_escrit: str = ""
    dt_ini: str = ""
    dt_fin: str = ""
    nome: str = ""
    cnpj: str = ""
    uf: str = ""
    cod_mun: str = ""


@dataclass
class ProductInfo:
    cod_item: str
    descr_item: str
    cod_ncm: str
    aliq_icms: Optional[Decimal] = None


@dataclass
class C870Record:
    line_number: int
    cod_item: str
    cfop: str
    vl_item: Decimal
    vl_desc: Decimal
    cst_pis: str
    vl_bc_pis: Decimal
    aliq_pis: Decimal
    vl_pis: Decimal
    cst_cofins: str
    vl_bc_cofins: Decimal
    aliq_cofins: Decimal
    vl_cofins: Decimal
    cod_cta: str
    raw_line: str


@dataclass
class CalculationResult:
    line_number: int
    cod_item: str
    ncm: str
    cfop: str
    vl_item: Decimal
    vl_bc_pis_orig: Decimal
    vl_pis_orig: Decimal
    vl_bc_cofins_orig: Decimal
    vl_cofins_orig: Decimal
    mva: Decimal
    aliq_icms: Decimal
    base_icms_st: Decimal
    valor_icms_st: Decimal
    vl_bc_pis_new: Decimal
    vl_pis_new: Decimal
    vl_bc_cofins_new: Decimal
    vl_cofins_new: Decimal
    economia_pis: Decimal
    economia_cofins: Decimal
    economia_total: Decimal
    status: str
    skip_reason: Optional[str] = None


@dataclass
class MonthSummary:
    month: str
    year: str
    month_name: str
    total_records: int
    total_calculated: int
    total_skipped: int
    pis_original: Decimal
    pis_adjusted: Decimal
    pis_credit: Decimal
    cofins_original: Decimal
    cofins_adjusted: Decimal
    cofins_credit: Decimal
    total_credit: Decimal
    savings_percentage: Decimal


# =============================================================================
# CLASSES DE PROCESSAMENTO
# =============================================================================

class SpedParser:
    """Parser de arquivos SPED Contribuições"""
    
    def __init__(self):
        self.header: Optional[SpedHeader] = None
        self.products: Dict[str, ProductInfo] = {}
        self.lines: List[str] = []
        self.line_count = 0
        self.c870_count = 0
    
    def parse_decimal(self, value: str) -> Decimal:
        if not value or value.strip() == '':
            return Decimal('0')
        clean = value.strip().replace(',', '.')
        try:
            return Decimal(clean)
        except:
            return Decimal('0')
    
    def parse_header(self, fields: List[str]) -> SpedHeader:
        return SpedHeader(
            cod_ver=fields[1] if len(fields) > 1 else '',
            tipo_escrit=fields[2] if len(fields) > 2 else '',
            dt_ini=fields[5] if len(fields) > 5 else '',
            dt_fin=fields[6] if len(fields) > 6 else '',
            nome=fields[7] if len(fields) > 7 else '',
            cnpj=fields[8] if len(fields) > 8 else '',
            uf=fields[9] if len(fields) > 9 else '',
            cod_mun=fields[10] if len(fields) > 10 else ''
        )
    
    def parse_product(self, fields: List[str]) -> ProductInfo:
        aliq_str = fields[11] if len(fields) > 11 else ''
        ncm_raw = fields[7] if len(fields) > 7 else ''
        ncm = ncm_raw[:8] if ncm_raw else ''
        
        return ProductInfo(
            cod_item=fields[1] if len(fields) > 1 else '',
            descr_item=fields[2] if len(fields) > 2 else '',
            cod_ncm=ncm,
            aliq_icms=self.parse_decimal(aliq_str) if aliq_str else None
        )
    
    def parse_c870(self, line_number: int, fields: List[str], raw_line: str) -> C870Record:
        return C870Record(
            line_number=line_number,
            cod_item=fields[1] if len(fields) > 1 else '',
            cfop=fields[2] if len(fields) > 2 else '',
            vl_item=self.parse_decimal(fields[3]) if len(fields) > 3 else Decimal('0'),
            vl_desc=self.parse_decimal(fields[4]) if len(fields) > 4 else Decimal('0'),
            cst_pis=fields[5] if len(fields) > 5 else '',
            vl_bc_pis=self.parse_decimal(fields[6]) if len(fields) > 6 else Decimal('0'),
            aliq_pis=self.parse_decimal(fields[7]) if len(fields) > 7 else Decimal('0'),
            vl_pis=self.parse_decimal(fields[8]) if len(fields) > 8 else Decimal('0'),
            cst_cofins=fields[9] if len(fields) > 9 else '',
            vl_bc_cofins=self.parse_decimal(fields[10]) if len(fields) > 10 else Decimal('0'),
            aliq_cofins=self.parse_decimal(fields[11]) if len(fields) > 11 else Decimal('0'),
            vl_cofins=self.parse_decimal(fields[12]) if len(fields) > 12 else Decimal('0'),
            cod_cta=fields[13] if len(fields) > 13 else '',
            raw_line=raw_line
        )
    
    def load_content(self, content: str) -> None:
        self.lines = content.split('\n')
        self.line_count = len(self.lines)
        
        for line_num, line in enumerate(self.lines, 1):
            line = line.strip()
            if not line:
                continue
            
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            
            fields = line.split('|')
            record_type = fields[0] if fields else ''
            
            if record_type == '0000':
                self.header = self.parse_header(fields)
            elif record_type == '0200':
                product = self.parse_product(fields)
                self.products[product.cod_item] = product
            elif record_type == 'C870':
                self.c870_count += 1
    
    def get_c870_records(self) -> Generator[C870Record, None, None]:
        for line_num, line in enumerate(self.lines, 1):
            line = line.strip()
            if not line:
                continue
            
            original_line = line
            if line.startswith('|'):
                line = line[1:]
            if line.endswith('|'):
                line = line[:-1]
            
            fields = line.split('|')
            if fields[0] == 'C870':
                yield self.parse_c870(line_num, fields, original_line)
    
    def get_ncm_for_item(self, cod_item: str) -> Optional[str]:
        product = self.products.get(cod_item)
        return product.cod_ncm if product else None


class ProductBaseLoader:
    """Carrega base de produtos do cliente (Excel)"""
    
    def __init__(self):
        self.products_by_ncm: Dict[str, Dict] = {}
    
    def load_dataframe(self, df: pd.DataFrame) -> int:
        col_map = {}
        for col in df.columns:
            col_lower = str(col).lower()
            if col_lower == 'ncm':
                col_map['ncm'] = col
            elif col_lower == 'capitulo':
                col_map['capitulo'] = col
            elif col_lower == 'item':
                col_map['item'] = col
            elif 'mva' in col_lower or 'iva' in col_lower:
                if 'ajust' in col_lower:
                    col_map['mva_adjusted'] = col
                elif 'import' not in col_lower:
                    col_map['mva'] = col
            elif 'aliq' in col_lower and 'entrada' in col_lower:
                col_map['aliq_icms'] = col
        
        count = 0
        for _, row in df.iterrows():
            ncm = None
            if 'ncm' in col_map and pd.notna(row[col_map['ncm']]):
                ncm_raw = str(row[col_map['ncm']]).strip()
                if ncm_raw and ncm_raw not in ['', 'nan']:
                    ncm = ncm_raw.replace('.', '').replace('-', '').zfill(8)[:8]
            
            if not ncm and 'capitulo' in col_map and 'item' in col_map:
                try:
                    cap = str(int(float(row[col_map['capitulo']]))).zfill(4)
                    item = str(int(float(row[col_map['item']]))).zfill(4)
                    ncm = cap + item
                except:
                    continue
            
            if not ncm or len(ncm) != 8:
                continue
            
            mva = None
            if 'mva' in col_map and pd.notna(row[col_map['mva']]):
                try:
                    mva_str = str(row[col_map['mva']]).replace(',', '.').replace('%', '')
                    mva = Decimal(mva_str)
                except:
                    pass
            
            aliq = Decimal('18')
            if 'aliq_icms' in col_map and pd.notna(row[col_map['aliq_icms']]):
                try:
                    aliq_str = str(row[col_map['aliq_icms']]).replace(',', '.').replace('%', '')
                    aliq = Decimal(aliq_str)
                except:
                    pass
            
            if mva is not None:
                self.products_by_ncm[ncm] = {
                    'ncm': ncm,
                    'mva': mva,
                    'aliq_icms': aliq
                }
                count += 1
        
        return count
    
    def get_product_by_ncm(self, ncm: str) -> Optional[Dict]:
        return self.products_by_ncm.get(ncm)


class IcmsStCalculator:
    """Calculadora de exclusão ICMS-ST"""
    
    def __init__(self, product_base: ProductBaseLoader, cfops_elegiveis: set):
        self.product_base = product_base
        self.cfops_elegiveis = cfops_elegiveis
    
    def calculate(self, record: C870Record, ncm: Optional[str]) -> CalculationResult:
        base = CalculationResult(
            line_number=record.line_number,
            cod_item=record.cod_item,
            ncm=ncm or '',
            cfop=record.cfop,
            vl_item=record.vl_item,
            vl_bc_pis_orig=record.vl_bc_pis,
            vl_pis_orig=record.vl_pis,
            vl_bc_cofins_orig=record.vl_bc_cofins,
            vl_cofins_orig=record.vl_cofins,
            mva=Decimal('0'),
            aliq_icms=Decimal('0'),
            base_icms_st=Decimal('0'),
            valor_icms_st=Decimal('0'),
            vl_bc_pis_new=record.vl_bc_pis,
            vl_pis_new=record.vl_pis,
            vl_bc_cofins_new=record.vl_bc_cofins,
            vl_cofins_new=record.vl_cofins,
            economia_pis=Decimal('0'),
            economia_cofins=Decimal('0'),
            economia_total=Decimal('0'),
            status='skipped'
        )
        
        if record.cfop not in self.cfops_elegiveis:
            base.skip_reason = f'CFOP {record.cfop} não elegível'
            return base
        
        if not ncm:
            base.skip_reason = 'NCM não encontrado'
            return base
        
        product = self.product_base.get_product_by_ncm(ncm)
        if not product:
            base.skip_reason = 'NCM sem MVA na base'
            return base
        
        mva = product['mva']
        aliq_icms = product.get('aliq_icms', Decimal('18'))

        if mva <= 0:
            base.skip_reason = 'MVA zero ou negativo'
            return base

        # Passo 4: Cálculo da exclusão ICMS-ST
        # Fórmula: VL_BC - ((VL_BC * MVA%) * ALIQ_ICMS%)
        mva_decimal = mva / Decimal('100')
        aliq_icms_decimal = aliq_icms / Decimal('100')

        # Exclusão para PIS (baseado no campo 7 - VL_BC_PIS)
        exclusao_pis = record.vl_bc_pis * mva_decimal * aliq_icms_decimal
        vl_bc_pis_new = (record.vl_bc_pis - exclusao_pis).quantize(Decimal('0.01'), ROUND_HALF_UP)
        if vl_bc_pis_new < 0:
            vl_bc_pis_new = Decimal('0')

        # Exclusão para COFINS (baseado no campo 11 - VL_BC_COFINS)
        exclusao_cofins = record.vl_bc_cofins * mva_decimal * aliq_icms_decimal
        vl_bc_cofins_new = (record.vl_bc_cofins - exclusao_cofins).quantize(Decimal('0.01'), ROUND_HALF_UP)
        if vl_bc_cofins_new < 0:
            vl_bc_cofins_new = Decimal('0')

        # Passo 6: Novos valores = BC_nova * alíquota (campos 8 e 12 têm 4 casas decimais)
        vl_pis_new = (vl_bc_pis_new * record.aliq_pis / Decimal('100')).quantize(Decimal('0.01'), ROUND_HALF_UP)
        vl_cofins_new = (vl_bc_cofins_new * record.aliq_cofins / Decimal('100')).quantize(Decimal('0.01'), ROUND_HALF_UP)

        # Valores para relatório
        # base_icms_st = VL_BC_PIS * MVA% (base intermediária do cálculo)
        base_icms_st = (record.vl_bc_pis * mva_decimal).quantize(Decimal('0.01'), ROUND_HALF_UP)
        valor_icms_st = exclusao_pis.quantize(Decimal('0.01'), ROUND_HALF_UP)

        economia_pis = record.vl_pis - vl_pis_new
        economia_cofins = record.vl_cofins - vl_cofins_new

        return CalculationResult(
            line_number=record.line_number,
            cod_item=record.cod_item,
            ncm=ncm,
            cfop=record.cfop,
            vl_item=record.vl_item,
            vl_bc_pis_orig=record.vl_bc_pis,
            vl_pis_orig=record.vl_pis,
            vl_bc_cofins_orig=record.vl_bc_cofins,
            vl_cofins_orig=record.vl_cofins,
            mva=mva,
            aliq_icms=aliq_icms,
            base_icms_st=base_icms_st,
            valor_icms_st=valor_icms_st,
            vl_bc_pis_new=vl_bc_pis_new,
            vl_pis_new=vl_pis_new,
            vl_bc_cofins_new=vl_bc_cofins_new,
            vl_cofins_new=vl_cofins_new,
            economia_pis=economia_pis,
            economia_cofins=economia_cofins,
            economia_total=economia_pis + economia_cofins,
            status='calculated'
        )


class SpedWriter:
    """Gera arquivo SPED retificado"""
    
    def __init__(self, parser: SpedParser, results: List[CalculationResult]):
        self.parser = parser
        self.results_by_line = {r.line_number: r for r in results if r.status == 'calculated'}
    
    def format_decimal(self, value: Decimal) -> str:
        return str(value.quantize(Decimal('0.01'))).replace('.', ',')
    
    def generate(self) -> str:
        modified_lines = []
        
        for line_num, line in enumerate(self.parser.lines, 1):
            original = line.strip()
            
            if not original:
                modified_lines.append(line)
                continue
            
            result = self.results_by_line.get(line_num)
            if not result:
                modified_lines.append(line)
                continue
            
            temp = original
            if temp.startswith('|'):
                temp = temp[1:]
            if temp.endswith('|'):
                temp = temp[:-1]
            
            fields = temp.split('|')
            
            if len(fields) >= 13:
                fields[6] = self.format_decimal(result.vl_bc_pis_new)
                fields[8] = self.format_decimal(result.vl_pis_new)
                fields[10] = self.format_decimal(result.vl_bc_cofins_new)
                fields[12] = self.format_decimal(result.vl_cofins_new)
                
                new_line = '|' + '|'.join(fields) + '|'
                modified_lines.append(new_line)
            else:
                modified_lines.append(line)
        
        return '\n'.join(modified_lines)


# =============================================================================
# GERADORES DE OUTPUT
# =============================================================================

def generate_excel(all_results: Dict[str, List[CalculationResult]], summaries: List[MonthSummary]) -> bytes:
    """Gera Excel consolidado com uma aba por mês"""
    wb = Workbook()
    wb.remove(wb.active)
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2E7D32')
    money_fill = PatternFill('solid', fgColor='E8F5E9')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aba de resumo
    ws = wb.create_sheet(title='RESUMO', index=0)
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'RESUMO CONSOLIDADO - EXCLUSÃO ICMS-ST DA BASE PIS/COFINS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['Mês/Ano', 'Registros', 'Calculados', 'PIS Original', 'PIS Crédito', 
               'COFINS Original', 'COFINS Crédito', 'Crédito Total']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, summary in enumerate(summaries, 4):
        ws.cell(row=row_num, column=1, value=f'{summary.month_name}/{summary.year}')
        ws.cell(row=row_num, column=2, value=summary.total_records)
        ws.cell(row=row_num, column=3, value=summary.total_calculated)
        ws.cell(row=row_num, column=4, value=float(summary.pis_original))
        ws.cell(row=row_num, column=5, value=float(summary.pis_credit))
        ws.cell(row=row_num, column=6, value=float(summary.cofins_original))
        ws.cell(row=row_num, column=7, value=float(summary.cofins_credit))
        ws.cell(row=row_num, column=8, value=float(summary.total_credit))
        
        for col in range(1, 9):
            ws.cell(row=row_num, column=col).border = thin_border
    
    total_row = len(summaries) + 4
    ws.cell(row=total_row, column=1, value='TOTAL')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    for col in [2, 3, 4, 5, 6, 7, 8]:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({col_letter}4:{col_letter}{total_row-1})'
        cell.font = Font(bold=True)
        cell.fill = money_fill
        cell.border = thin_border
    
    for row in range(4, total_row + 1):
        for col in [4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'
    
    ws.column_dimensions['A'].width = 15
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Abas por mês
    for sheet_name, results in all_results.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        calculated = [r for r in results if r.status == 'calculated']
        
        headers = [
            'Linha', 'Cod Item', 'NCM', 'CFOP', 'Valor Item',
            'BC PIS Orig', 'PIS Orig', 'BC COFINS Orig', 'COFINS Orig',
            'MVA %', 'ICMS-ST', 'BC PIS Nova', 'PIS Novo',
            'BC COFINS Nova', 'COFINS Novo', 'Economia PIS', 'Economia COFINS', 'Economia Total'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, result in enumerate(calculated, 2):
            ws.cell(row=row_num, column=1, value=result.line_number)
            ws.cell(row=row_num, column=2, value=result.cod_item)
            ws.cell(row=row_num, column=3, value=result.ncm)
            ws.cell(row=row_num, column=4, value=result.cfop)
            ws.cell(row=row_num, column=5, value=float(result.vl_item))
            ws.cell(row=row_num, column=6, value=float(result.vl_bc_pis_orig))
            ws.cell(row=row_num, column=7, value=float(result.vl_pis_orig))
            ws.cell(row=row_num, column=8, value=float(result.vl_bc_cofins_orig))
            ws.cell(row=row_num, column=9, value=float(result.vl_cofins_orig))
            ws.cell(row=row_num, column=10, value=float(result.mva))
            ws.cell(row=row_num, column=11, value=float(result.valor_icms_st))
            ws.cell(row=row_num, column=12, value=float(result.vl_bc_pis_new))
            ws.cell(row=row_num, column=13, value=float(result.vl_pis_new))
            ws.cell(row=row_num, column=14, value=float(result.vl_bc_cofins_new))
            ws.cell(row=row_num, column=15, value=float(result.vl_cofins_new))
            ws.cell(row=row_num, column=16, value=float(result.economia_pis))
            ws.cell(row=row_num, column=17, value=float(result.economia_cofins))
            ws.cell(row=row_num, column=18, value=float(result.economia_total))
            
            for col in range(1, 19):
                ws.cell(row=row_num, column=col).border = thin_border
        
        for col in range(1, 19):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf(summaries: List[MonthSummary], company_name: str, cnpj: str) -> bytes:
    """Gera relatório PDF consolidado"""
    output = io.BytesIO()
    
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=20))
    styles.add(ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, textColor=colors.grey, spaceAfter=20))
    
    elements = []
    
    elements.append(Paragraph('RELATÓRIO DE RECUPERAÇÃO DE CRÉDITOS TRIBUTÁRIOS', styles['Title2']))
    elements.append(Paragraph('Exclusão do ICMS-ST da Base de Cálculo de PIS/COFINS', styles['Subtitle']))
    
    elements.append(Paragraph(f'<b>Empresa:</b> {company_name}', styles['Normal']))
    elements.append(Paragraph(f'<b>CNPJ:</b> {cnpj}', styles['Normal']))
    elements.append(Paragraph(f'<b>Data do Relatório:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 20))
    
    total_credit = sum(s.total_credit for s in summaries)
    total_records = sum(s.total_records for s in summaries)
    total_calculated = sum(s.total_calculated for s in summaries)
    
    elements.append(Paragraph('<b>RESUMO EXECUTIVO</b>', styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    summary_data = [
        ['Período Analisado', f'{summaries[0].month_name}/{summaries[0].year} a {summaries[-1].month_name}/{summaries[-1].year}'],
        ['Total de Registros Processados', f'{total_records:,}'],
        ['Registros com Cálculo Aplicado', f'{total_calculated:,}'],
        ['Taxa de Aproveitamento', f'{(total_calculated/total_records*100):.1f}%'],
        ['CRÉDITO TOTAL RECUPERÁVEL', f'R$ {float(total_credit):,.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[250, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.18, 0.49, 0.20)),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    elements.append(Paragraph('<b>DETALHAMENTO MENSAL</b>', styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    monthly_data = [['Mês/Ano', 'Registros', 'Calculados', 'Crédito PIS', 'Crédito COFINS', 'Crédito Total']]
    
    for s in summaries:
        monthly_data.append([
            f'{s.month_name}/{s.year}',
            f'{s.total_records:,}',
            f'{s.total_calculated:,}',
            f'R$ {float(s.pis_credit):,.2f}',
            f'R$ {float(s.cofins_credit):,.2f}',
            f'R$ {float(s.total_credit):,.2f}'
        ])
    
    monthly_data.append([
        'TOTAL',
        f'{total_records:,}',
        f'{total_calculated:,}',
        f'R$ {sum(float(s.pis_credit) for s in summaries):,.2f}',
        f'R$ {sum(float(s.cofins_credit) for s in summaries):,.2f}',
        f'R$ {float(total_credit):,.2f}'
    ])
    
    monthly_table = Table(monthly_data, colWidths=[80, 70, 70, 90, 90, 90])
    monthly_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.18, 0.49, 0.20)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(monthly_table)
    
    doc.build(elements)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================

MONTH_NAMES = {
    '01': 'Janeiro', '02': 'Fevereiro', '03': 'Março',
    '04': 'Abril', '05': 'Maio', '06': 'Junho',
    '07': 'Julho', '08': 'Agosto', '09': 'Setembro',
    '10': 'Outubro', '11': 'Novembro', '12': 'Dezembro'
}


def extract_month_year_from_filename(filename: str) -> Tuple[str, str]:
    """Extrai mês e ano do nome do arquivo"""
    match = re.search(r'(\d{2})[-_]?(\d{4})', filename)
    if match:
        return match.group(1), match.group(2)
    return '', ''


def extract_month_year_from_sped_date(dt_ini: str) -> Tuple[str, str]:
    """Extrai mês e ano da data do SPED (formato DDMMYYYY)"""
    if dt_ini and len(dt_ini) >= 8:
        month = dt_ini[2:4]
        year = dt_ini[4:8]
        return month, year
    return '', ''


def extract_month_year(filename: str, header: Optional[SpedHeader] = None) -> Tuple[str, str]:
    """Extrai mês e ano, priorizando o header do SPED"""
    # Primeiro tenta extrair do header do SPED (mais confiável)
    if header and header.dt_ini:
        month, year = extract_month_year_from_sped_date(header.dt_ini)
        if month and year:
            return month, year

    # Fallback: extrair do nome do arquivo
    month, year = extract_month_year_from_filename(filename)
    if month and year:
        return month, year

    return '00', '0000'


def main():
    # Header
    st.markdown("""
    <div class="header-container">
        <h1 class="header-title">🧾 OmniAI Fiscal</h1>
        <p class="header-subtitle">Exclusão do ICMS-ST da Base de Cálculo PIS/COFINS</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("### ⚙️ Configurações")
        
        st.markdown("#### CFOPs Elegíveis")
        cfop_5405 = st.checkbox("5405 - Venda ST Substituído", value=True)
        cfop_5403 = st.checkbox("5403 - Venda ST Substituto", value=False)
        cfop_5401 = st.checkbox("5401 - Venda Produção ST", value=False)
        cfop_5102 = st.checkbox("5102 - Venda Revenda", value=False)
        
        cfops_selecionados = set()
        if cfop_5405:
            cfops_selecionados.add('5405')
        if cfop_5403:
            cfops_selecionados.add('5403')
        if cfop_5401:
            cfops_selecionados.add('5401')
        if cfop_5102:
            cfops_selecionados.add('5102')
        
        st.markdown("---")
        
        st.markdown("#### 📊 Sobre")
        st.info("""
        Esta ferramenta calcula os créditos de PIS/COFINS 
        decorrentes da exclusão do ICMS-ST da base de cálculo,
        conforme entendimento jurisprudencial.
        """)
        
        st.markdown("#### 📋 Metodologia")
        st.markdown("""
        1. **Identificação**: Registros C870 com CFOPs selecionados
        2. **Enriquecimento**: NCM → MVA da base de produtos
        3. **Cálculo**: Base ICMS-ST = Valor × (1 + MVA%)
        4. **Exclusão**: Nova BC = BC Original - ICMS-ST
        5. **Crédito**: Diferença dos tributos
        """)
    
    # Main content
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📁 Base de Produtos")
        st.markdown("Upload do arquivo Excel com NCMs e MVAs")
        produto_file = st.file_uploader(
            "Arraste ou clique para upload",
            type=['xlsx', 'xls'],
            key='produtos',
            help="Arquivo Excel com colunas: NCM (ou Capitulo+Item), MVA, Alíquota ICMS"
        )
        
        if produto_file:
            st.success(f"✅ {produto_file.name}")
    
    with col2:
        st.markdown("### 📄 Arquivos SPED")
        st.markdown("Upload dos arquivos SPED Contribuições (.txt)")
        sped_files = st.file_uploader(
            "Arraste ou clique para upload",
            type=['txt'],
            accept_multiple_files=True,
            key='sped',
            help="Arquivos SPED Contribuições no formato TXT"
        )
        
        if sped_files:
            st.success(f"✅ {len(sped_files)} arquivo(s) selecionado(s)")
    
    st.markdown("---")
    
    # Botão de processamento
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
    
    with col_btn2:
        process_btn = st.button(
            "🚀 PROCESSAR ARQUIVOS",
            type="primary",
            use_container_width=True,
            disabled=not (produto_file and sped_files and cfops_selecionados)
        )
    
    if not cfops_selecionados:
        st.warning("⚠️ Selecione pelo menos um CFOP elegível na barra lateral.")
    
    # Processamento
    if process_btn and produto_file and sped_files:
        
        with st.spinner("Carregando base de produtos..."):
            df_produtos = pd.read_excel(produto_file)
            product_base = ProductBaseLoader()
            ncm_count = product_base.load_dataframe(df_produtos)
        
        st.info(f"📊 Base carregada: **{ncm_count:,}** NCMs com MVA")
        
        # Ordenar arquivos por data
        sorted_files = sorted(sped_files, key=lambda f: extract_month_year_from_filename(f.name))
        
        summaries: List[MonthSummary] = []
        all_results: Dict[str, List[CalculationResult]] = {}
        sped_outputs: Dict[str, str] = {}
        company_name = ""
        cnpj = ""
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        calculator = IcmsStCalculator(product_base, cfops_selecionados)
        
        for idx, sped_file in enumerate(sorted_files):
            status_text.text(f"Processando arquivo {idx + 1} de {len(sorted_files)}...")

            # Parse SPED
            content = sped_file.read().decode('latin-1')
            parser = SpedParser()
            parser.load_content(content)

            # Extrair mês/ano do header do SPED (mais confiável que o nome do arquivo)
            month, year = extract_month_year(sped_file.name, parser.header)
            month_name = MONTH_NAMES.get(month, month)

            status_text.text(f"Processando {month_name}/{year}...")

            if parser.header and not company_name:
                company_name = parser.header.nome
                cnpj = parser.header.cnpj
            
            # Calcular
            results: List[CalculationResult] = []
            for record in parser.get_c870_records():
                ncm = parser.get_ncm_for_item(record.cod_item)
                result = calculator.calculate(record, ncm)
                results.append(result)
            
            # Estatísticas
            calculated = [r for r in results if r.status == 'calculated']
            skipped = [r for r in results if r.status == 'skipped']
            
            pis_orig = sum(r.vl_pis_orig for r in calculated)
            pis_new = sum(r.vl_pis_new for r in calculated)
            cofins_orig = sum(r.vl_cofins_orig for r in calculated)
            cofins_new = sum(r.vl_cofins_new for r in calculated)
            
            pis_credit = pis_orig - pis_new
            cofins_credit = cofins_orig - cofins_new
            total_credit = pis_credit + cofins_credit
            
            total_original = pis_orig + cofins_orig
            savings_pct = (total_credit / total_original * 100) if total_original > 0 else Decimal('0')
            
            summary = MonthSummary(
                month=month,
                year=year,
                month_name=month_name,
                total_records=len(results),
                total_calculated=len(calculated),
                total_skipped=len(skipped),
                pis_original=pis_orig,
                pis_adjusted=pis_new,
                pis_credit=pis_credit,
                cofins_original=cofins_orig,
                cofins_adjusted=cofins_new,
                cofins_credit=cofins_credit,
                total_credit=total_credit,
                savings_percentage=savings_pct.quantize(Decimal('0.01'), ROUND_HALF_UP) if isinstance(savings_pct, Decimal) else Decimal('0')
            )
            
            summaries.append(summary)
            
            sheet_name = f'{month_name[:3]}_{year}'
            all_results[sheet_name] = results
            
            # Gerar SPED retificado
            writer = SpedWriter(parser, results)
            sped_outputs[f'SPED_RETIFICADO_{month}_{year}.txt'] = writer.generate()
            
            progress_bar.progress((idx + 1) / len(sorted_files))
        
        status_text.text("✅ Processamento concluído!")
        progress_bar.progress(1.0)
        
        st.markdown("---")
        
        # Resultados
        st.markdown("## 📊 Resultados")
        
        # Métricas principais
        total_credit = sum(s.total_credit for s in summaries)
        total_records = sum(s.total_records for s in summaries)
        total_calculated = sum(s.total_calculated for s in summaries)
        total_pis = sum(s.pis_credit for s in summaries)
        total_cofins = sum(s.cofins_credit for s in summaries)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                label="💰 Crédito Total",
                value=f"R$ {float(total_credit):,.2f}",
                delta=f"{(total_calculated/total_records*100):.1f}% aproveitamento"
            )
        
        with col2:
            st.metric(
                label="📄 Registros",
                value=f"{total_records:,}",
                delta=f"{total_calculated:,} calculados"
            )
        
        with col3:
            st.metric(
                label="🔵 Crédito PIS",
                value=f"R$ {float(total_pis):,.2f}"
            )
        
        with col4:
            st.metric(
                label="🟢 Crédito COFINS",
                value=f"R$ {float(total_cofins):,.2f}"
            )
        
        # Empresa
        st.markdown(f"""
        **Empresa:** {company_name}  
        **CNPJ:** {cnpj}  
        **Período:** {summaries[0].month_name}/{summaries[0].year} a {summaries[-1].month_name}/{summaries[-1].year}
        """)
        
        # Tabela detalhada
        st.markdown("### 📅 Detalhamento Mensal")
        
        df_summary = pd.DataFrame([{
            'Mês/Ano': f'{s.month_name}/{s.year}',
            'Registros': s.total_records,
            'Calculados': s.total_calculated,
            'Crédito PIS': f'R$ {float(s.pis_credit):,.2f}',
            'Crédito COFINS': f'R$ {float(s.cofins_credit):,.2f}',
            'Crédito Total': f'R$ {float(s.total_credit):,.2f}'
        } for s in summaries])
        
        st.dataframe(df_summary, use_container_width=True, hide_index=True)
        
        # Downloads
        st.markdown("### 📥 Downloads")

        # Chave única para identificar este processamento (baseada nos arquivos e CFOPs)
        cache_key = f"downloads_{hash(tuple(sorted([f.name for f in sorted_files])) + tuple(sorted(cfops_selecionados)))}"

        # Verificar se já temos os arquivos em cache ou se precisa gerar
        if cache_key not in st.session_state:
            with st.spinner("Gerando arquivos para download..."):
                # Excel
                excel_data = generate_excel(all_results, summaries)

                # PDF
                pdf_data = generate_pdf(summaries, company_name, cnpj)

                # ZIP com SPEDs
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for filename, content in sped_outputs.items():
                        zip_file.writestr(filename, content.encode('latin-1'))
                zip_buffer.seek(0)
                zip_data = zip_buffer.getvalue()

                # JSON para API
                json_data = {
                    'empresa': company_name,
                    'cnpj': cnpj,
                    'periodo': f'{summaries[0].month_name}/{summaries[0].year} a {summaries[-1].month_name}/{summaries[-1].year}',
                    'processado_em': datetime.now().isoformat(),
                    'cfops_utilizados': list(cfops_selecionados),
                    'total_registros': total_records,
                    'total_calculados': total_calculated,
                    'credito_pis': float(total_pis),
                    'credito_cofins': float(total_cofins),
                    'credito_total': float(total_credit),
                    'meses': [
                        {
                            'mes': s.month_name,
                            'ano': s.year,
                            'registros': s.total_records,
                            'calculados': s.total_calculated,
                            'credito_pis': float(s.pis_credit),
                            'credito_cofins': float(s.cofins_credit),
                            'credito_total': float(s.total_credit)
                        }
                        for s in summaries
                    ]
                }
                json_str = json.dumps(json_data, ensure_ascii=False, indent=2)

                # ZIP completo com todos os arquivos
                primeiro_arquivo = sorted_files[0].name
                nome_base = primeiro_arquivo.rsplit('.', 1)[0] if '.' in primeiro_arquivo else primeiro_arquivo
                nome_zip = f"{nome_base}.zip"

                all_files_buffer = io.BytesIO()
                with zipfile.ZipFile(all_files_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_all:
                    zip_all.writestr("DE_PARA_CONSOLIDADO.xlsx", excel_data)
                    zip_all.writestr("RELATORIO_CONSOLIDADO.pdf", pdf_data)
                    zip_all.writestr("resumo_consolidado.json", json_str)
                    for filename, content in sped_outputs.items():
                        zip_all.writestr(f"SPEDS_RETIFICADOS/{filename}", content.encode('latin-1'))
                all_files_buffer.seek(0)
                all_files_data = all_files_buffer.getvalue()

                # Armazenar no session_state
                st.session_state[cache_key] = {
                    'excel_data': excel_data,
                    'pdf_data': pdf_data,
                    'zip_data': zip_data,
                    'json_data': json_data,
                    'json_str': json_str,
                    'all_files_data': all_files_data,
                    'nome_zip': nome_zip
                }

            st.success("Arquivos gerados com sucesso!")
        else:
            # Recuperar do cache
            cached = st.session_state[cache_key]
            excel_data = cached['excel_data']
            pdf_data = cached['pdf_data']
            zip_data = cached['zip_data']
            json_data = cached['json_data']
            json_str = cached['json_str']
            all_files_data = cached['all_files_data']
            nome_zip = cached['nome_zip']

        col_dl1, col_dl2, col_dl3 = st.columns(3)

        with col_dl1:
            st.download_button(
                label="📊 Download Excel (De/Para)",
                data=excel_data,
                file_name="DE_PARA_CONSOLIDADO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        with col_dl2:
            st.download_button(
                label="📄 Download PDF (Relatório)",
                data=pdf_data,
                file_name="RELATORIO_CONSOLIDADO.pdf",
                mime="application/pdf",
                use_container_width=True
            )

        with col_dl3:
            st.download_button(
                label="📦 Download SPEDs Retificados",
                data=zip_data,
                file_name="SPEDS_RETIFICADOS.zip",
                mime="application/zip",
                use_container_width=True
            )

        with st.expander("🔧 JSON para Integração"):
            st.json(json_data)
            st.download_button(
                label="Download JSON",
                data=json_str,
                file_name="resumo_consolidado.json",
                mime="application/json"
            )

        # Download de todos os arquivos em um único ZIP
        st.markdown("---")
        st.markdown("### 📦 Download Completo")

        st.download_button(
            label="⬇️ Download Todos os Arquivos (ZIP)",
            data=all_files_data,
            file_name=nome_zip,
            mime="application/zip",
            use_container_width=True,
            type="primary"
        )


if __name__ == '__main__':
    main()
